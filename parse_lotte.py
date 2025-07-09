from openpyxl import Workbook, load_workbook
import xlrd
import json
from pathlib import Path
from datetime import datetime

# Load config
with open("config/column_map_lotte.json", "r", encoding="utf-8") as f:
    config = json.load(f)["lotte_excel"]

with open("config/paths.json", "r", encoding="utf-8") as f:
    paths = json.load(f)["paths"]

input_dir = Path(paths["input_lotte"])
output_dir = Path(paths["output_lotte"])
output_dir.mkdir(parents=True, exist_ok=True)
log_path = output_dir / "parse_lotte.log"

field_map = config["field_map"]
date_fmt = config["date_format"]

logs = []

def convert_xls_to_workbook(xls_path):
    print(f"[DEBUG] Converting XLS: {xls_path.name}")
    book = xlrd.open_workbook(xls_path)
    sheet = book.sheet_by_index(0)
    wb = Workbook()
    ws = wb.active
    for row_idx in range(sheet.nrows):
        ws.append(sheet.row_values(row_idx))
    return wb

def parse_lotte_workbook(wb, file_name):
    print(f"[DEBUG] Parsing workbook: {file_name}")
    ws = wb.active
    order_blocks = []
    current_slip = None
    current_date = None
    current_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        print(f"[DEBUG] Row {row_idx}: {row}")
        slip_val = row[4]      # column E
        product_val = row[20]  # column U

        if slip_val and isinstance(slip_val, str) and "-" in slip_val:
            if current_slip and current_rows:
                order_blocks.append({
                    "delivery_date": current_date,
                    "source_file": file_name,
                    "order_slip": current_slip,
                    "rows": current_rows
                })
                print(f"[DEBUG] Block saved: {current_slip} with {len(current_rows)} rows")

            current_slip = slip_val.strip()
            current_rows = []
            current_date = None

            try:
                raw_date = row[12]
                if raw_date:
                    raw_date_str = str(raw_date).strip()
                    dt = datetime.strptime(raw_date_str, date_fmt)
                    current_date = dt.strftime("%Y-%m-%d")
                    print(f"[DEBUG] Date detected: {current_date}")
                else:
                    print(f"[DEBUG] No delivery date in row {row_idx}")
            except Exception as e:
                print(f"[WARN] Failed to parse delivery date in slip header (row {row_idx}): {raw_date} — {e}")

            print(f"[DEBUG] New slip: {current_slip}")

        if not product_val or not current_slip:
            continue

        try:
            product_name = product_val
            unit_price = float(str(row[26]).replace(",", ""))   # column AA
            tax = float(str(row[27]).replace(",", ""))          # column AB
            qty = int(str(row[28]).replace(",", ""))            # column AC

            current_rows.append({
                "product_name": product_name,
                "qty": qty,
                "unit_price": unit_price,
                "tax": tax
            })
            print(f"[DEBUG] Row added: {product_name}, {qty}, {unit_price}, {tax}")

        except Exception as e:
            print(f"[ERROR] Row parse failed: {e}")
            continue

    if current_slip and current_rows:
        order_blocks.append({
            "delivery_date": current_date,
            "source_file": file_name,
            "order_slip": current_slip,
            "rows": current_rows
        })
        print(f"[DEBUG] Final block saved: {current_slip} with {len(current_rows)} rows")

    for block in order_blocks:
        out_file = output_dir / f"{file_name}__{block['order_slip']}.json"
        with open(out_file, "w", encoding="utf-8") as f:
            json.dump(block, f, ensure_ascii=False, indent=2)
        logs.append(f"[OK] {out_file.name} → {len(block['rows'])} rows")

# Main execution
for file in input_dir.glob("*"):
    print(f"[DEBUG] Processing file: {file.name}")
    try:
        if file.suffix == ".xls":
            wb = convert_xls_to_workbook(file)
        elif file.suffix == ".xlsx":
            wb = load_workbook(file, data_only=True)
        else:
            print(f"[SKIP] Unsupported file: {file.name}")
            continue

        parse_lotte_workbook(wb, file.name)

    except Exception as e:
        logs.append(f"[ERROR] {file.name}: {e}")
        print(f"[ERROR] Exception while processing {file.name}: {e}")

with open(log_path, "w", encoding="utf-8") as f:
    for line in logs:
        f.write(line + "\n")

for line in logs[-10:]:
    print(line)
